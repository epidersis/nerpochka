from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from sqlalchemy import text

from app.database import engine

REPORT_YEAR = int(os.getenv("REPORT_YEAR", "2025"))

SECTION_SHEETS = [
    {
        "section": "КИК",
        "sheet": "Раздел 1 (КИК)",
        "title": "Раздел 1. КИК",
        "note": "КЦСР=*****975**",
        "name_header": "Наименование мероприятия",
        "mode": "kcsr",
        "pattern": "%975%",
        "last_col": 45,
    },
    {
        "section": "СКК",
        "sheet": "Раздел 2 (СКК)",
        "title": "Раздел 2. Сведения о кассовых операциях за счет средств специальных казначейских кредитов (СКК)",
        "note": "КЦСР=*****970**",
        "name_header": "Наименование мероприятия",
        "mode": "kcsr",
        "pattern": "%970%",
        "last_col": 45,
    },
    {
        "section": "Раздел 2/3",
        "sheet": "23",
        "title": "Раздел 3. 2/3",
        "note": "КЦСР=*****6105*",
        "name_header": "Наименование мероприятия",
        "mode": "kcsr",
        "pattern": "%6105%",
        "last_col": 49,
    },
    {
        "section": "ОКВ",
        "sheet": "ОКВ",
        "title": "Раздел 4. Объекты капитальных вложений",
        "note": "ДопКР не равен 0",
        "name_header": "Наименование объекта",
        "mode": "okv",
        "pattern": None,
        "last_col": 56,
    },
]

NUMBER_FORMAT = '#,##0.00'
DATE_FORMAT = 'DD.MM.YYYY'

AMOUNT_KEYS = (
    "limit_amount",
    "obligation_amount",
    "cash_amount",
    "cash_contract_amount",
    "cash_subsidy_amount",
    "cash_mbt_amount",
    "local_limit_amount",
    "local_obligation_amount",
    "local_cash_amount",
    "local_cash_contract_amount",
    "local_cash_subsidy_amount",
    "local_cash_mbt_amount",
    "contract_amount",
    "mbt_amount",
    "subsidy_amount",
    "buau_cash_amount",
)


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    return float(value or 0)


def _first_text(*values: Any) -> str | None:
    for value in values:
        if value not in (None, ""):
            return str(value)
    return None


def _kvr_digits_sql(alias: str) -> str:
    return f"regexp_replace(coalesce({alias}.kvr_code, ''), '[^0-9]', '', 'g')"


def _base_filter_sql(alias: str, mode: str) -> str:
    if mode == "okv":
        return f"nullif({alias}.kdr_code, '') is not null and {alias}.kdr_code not in ('0', '0.0')"
    return f"{alias}.kcsr_code ilike :pattern"


def _object_code_sql(alias: str, mode: str) -> str:
    if mode == "okv":
        return f"coalesce(nullif({alias}.kdr_code, ''), {alias}.kcsr_code)"
    return f"replace({alias}.kcsr_code, '.', '')"


def _object_name_sql(alias: str, mode: str) -> str:
    if mode == "okv":
        return f"coalesce(nullif({alias}.kdr_name, ''), nullif({alias}.estimate_name, ''), nullif({alias}.recipient_name, ''), nullif({alias}.kdr_code, ''), {alias}.kcsr_code)"
    return f"coalesce(nullif({alias}.estimate_name, ''), nullif({alias}.recipient_name, ''), {alias}.kcsr_code)"


def _is_subject_budget_sql(alias: str) -> str:
    return f"{alias}.budget_id ilike 'Областной бюджет%'"


def _is_local_budget_sql(alias: str) -> str:
    return f"not ({_is_subject_budget_sql(alias)})"


def _latest_rchb_period(conn):
    return conn.execute(
        text("""
            select max(period_to)
            from stg.budget_operations
            where amount is not null
              and extract(year from period_to) = :report_year
        """),
        {"report_year": REPORT_YEAR},
    ).scalar()


def _fetch_base_rows(conn, meta: dict[str, Any], kcsr_code: str | None) -> dict[str, dict[str, Any]]:
    mode = meta["mode"]
    latest_period = _latest_rchb_period(conn)
    params = {
        "pattern": meta["pattern"],
        "kcsr_code": kcsr_code,
        "latest_period": latest_period,
    }
    query = text(f"""
        select
            {_object_name_sql('b', mode)} as object_code,
            {_object_name_sql('b', mode)} as object_name,
            max(b.kcsr_code) as kcsr_code,
            max(b.kdr_code) as kdr_code,
            sum(b.amount) filter (where b.documentclass_id = '273' and {_is_subject_budget_sql('b')}) as limit_amount,
            sum(b.amount) filter (where b.documentclass_id = '313' and {_is_subject_budget_sql('b')}) as obligation_amount,
            sum(b.amount) filter (where b.documentclass_id = 'cash_execution' and {_is_subject_budget_sql('b')}) as cash_amount,
            sum(b.amount) filter (
                where b.documentclass_id = 'cash_execution'
                  and {_is_subject_budget_sql('b')}
                  and {_kvr_digits_sql('b')} like '2%%'
            ) as cash_contract_amount,
            sum(b.amount) filter (
                where b.documentclass_id = 'cash_execution'
                  and {_is_subject_budget_sql('b')}
                  and {_kvr_digits_sql('b')} like '6%%'
            ) as cash_subsidy_amount,
            sum(b.amount) filter (
                where b.documentclass_id = 'cash_execution'
                  and {_is_subject_budget_sql('b')}
                  and {_kvr_digits_sql('b')} like '5%%'
            ) as cash_mbt_amount,
            sum(b.amount) filter (where b.documentclass_id = '273' and {_is_local_budget_sql('b')}) as local_limit_amount,
            sum(b.amount) filter (where b.documentclass_id = '313' and {_is_local_budget_sql('b')}) as local_obligation_amount,
            sum(b.amount) filter (where b.documentclass_id = 'cash_execution' and {_is_local_budget_sql('b')}) as local_cash_amount,
            sum(b.amount) filter (
                where b.documentclass_id = 'cash_execution'
                  and {_is_local_budget_sql('b')}
                  and {_kvr_digits_sql('b')} like '2%%'
            ) as local_cash_contract_amount,
            sum(b.amount) filter (
                where b.documentclass_id = 'cash_execution'
                  and {_is_local_budget_sql('b')}
                  and {_kvr_digits_sql('b')} like '6%%'
            ) as local_cash_subsidy_amount,
            sum(b.amount) filter (
                where b.documentclass_id = 'cash_execution'
                  and {_is_local_budget_sql('b')}
                  and {_kvr_digits_sql('b')} like '5%%'
            ) as local_cash_mbt_amount
        from stg.budget_operations b
        where b.period_to = :latest_period
          and ({_base_filter_sql('b', mode)})
          and (:kcsr_code is null or replace(b.kcsr_code, '.', '') = replace(:kcsr_code, '.', ''))
        group by object_code, object_name
        order by object_name
    """)

    rows: dict[str, dict[str, Any]] = {}
    for row in conn.execute(query, params).mappings():
        rows[row["object_code"]] = {
            "object_code": row["object_code"],
            "object_name": row["object_name"],
            "kcsr_code": row["kcsr_code"],
            "kdr_code": row["kdr_code"],
            "limit_amount": _to_float(row["limit_amount"]),
            "obligation_amount": _to_float(row["obligation_amount"]),
                "cash_amount": _to_float(row["cash_amount"]),
                "cash_contract_amount": _to_float(row["cash_contract_amount"]),
                "cash_subsidy_amount": _to_float(row["cash_subsidy_amount"]),
                "cash_mbt_amount": _to_float(row["cash_mbt_amount"]),
                "local_limit_amount": _to_float(row["local_limit_amount"]),
                "local_obligation_amount": _to_float(row["local_obligation_amount"]),
                "local_cash_amount": _to_float(row["local_cash_amount"]),
                "local_cash_contract_amount": _to_float(row["local_cash_contract_amount"]),
                "local_cash_subsidy_amount": _to_float(row["local_cash_subsidy_amount"]),
                "local_cash_mbt_amount": _to_float(row["local_cash_mbt_amount"]),
            }
    return rows


def _merge_detail(target: dict[str, dict[str, Any]], key: str, prefix: str, row: dict[str, Any]) -> None:
    item = target.setdefault(key, {
        "object_code": key,
        "object_name": _first_text(row.get("object_name"), key),
        "limit_amount": 0.0,
        "obligation_amount": 0.0,
        "cash_amount": 0.0,
        "cash_contract_amount": 0.0,
        "cash_subsidy_amount": 0.0,
        "cash_mbt_amount": 0.0,
        "local_limit_amount": 0.0,
        "local_obligation_amount": 0.0,
        "local_cash_amount": 0.0,
        "local_cash_contract_amount": 0.0,
        "local_cash_subsidy_amount": 0.0,
        "local_cash_mbt_amount": 0.0,
    })
    if not item.get("object_name") or item["object_name"] == key:
        item["object_name"] = _first_text(row.get("object_name"), key)
    item[f"{prefix}_amount"] = item.get(f"{prefix}_amount", 0.0) + _to_float(row.get("amount"))
    item.setdefault(f"{prefix}_date", row.get("date_value"))
    item.setdefault(f"{prefix}_number", row.get("number_value"))
    item.setdefault(f"{prefix}_name", row.get("name_value"))


def _fetch_contracts(conn, target: dict[str, dict[str, Any]], meta: dict[str, Any], kcsr_code: str | None) -> None:
    mode = meta["mode"]
    params = {"pattern": meta["pattern"], "kcsr_code": kcsr_code}
    rchb_object_code = (
        "coalesce(nullif(rb.kdr_name, ''), nullif(rb.estimate_name, ''), nullif(b.kdr_code, ''), b.kcsr_code)"
        if mode == "okv"
        else "coalesce(nullif(rb.estimate_name, ''), b.kcsr_code)"
    )
    query = text(f"""
        with latest_rchb as (
            select max(period_to) as period_to
            from stg.budget_operations
            where amount is not null
              and extract(year from period_to) = :report_year
        )
        select
            {rchb_object_code} as object_code,
            {rchb_object_code} as object_name,
            c.con_date as date_value,
            c.con_number as number_value,
            coalesce(c.supplier_name, c.customer_name) as name_value,
            c.con_amount as amount
        from stg.gz_contracts c
        join stg.gz_budget_lines b on b.con_document_id = c.con_document_id
        left join stg.budget_operations rb
          on rb.period_to = (select period_to from latest_rchb)
         and replace(rb.kcsr_code, '.', '') = replace(b.kcsr_code, '.', '')
         and (
              :mode <> 'okv'
              or coalesce(rb.kdr_code, '') = coalesce(b.kdr_code, '')
         )
        where ({_base_filter_sql('b', mode)})
          and (:kcsr_code is null or replace(b.kcsr_code, '.', '') = replace(:kcsr_code, '.', ''))
          and c.con_amount is not null
    """)
    seen = set()
    for row in conn.execute(query, {**params, "report_year": REPORT_YEAR, "mode": mode}).mappings():
        dedupe_key = (row["object_code"], row["number_value"], row["amount"])
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        _merge_detail(target, row["object_code"], "contract", dict(row))


def _fetch_agreements(conn, target: dict[str, dict[str, Any]], meta: dict[str, Any], kcsr_code: str | None) -> None:
    mode = meta["mode"]
    params = {"pattern": meta["pattern"], "kcsr_code": kcsr_code}
    object_code_sql = "coalesce(nullif(a.kdr_code, ''), a.kcsr_code)" if mode == "okv" else "replace(a.kcsr_code, '.', '')"
    base_filter = (
        "nullif(a.kdr_code, '') is not null and a.kdr_code not in ('0', '0.0')"
        if mode == "okv"
        else "a.kcsr_code ilike :pattern"
    )
    query = text(f"""
        with latest_agreements as (
            select max(period_to) as period_to
            from stg.agreements
            where amount is not null
              and extract(year from period_to) = :report_year
        ),
        ranked as (
            select
                a.*,
                row_number() over (
                    partition by a.document_id, a.kcsr_code, coalesce(a.kdr_code, ''), a.documentclass_id
                    order by a.period_to desc, a.import_file_id desc
                ) as rn
            from stg.agreements a
            where ({base_filter})
              and (:kcsr_code is null or replace(a.kcsr_code, '.', '') = replace(:kcsr_code, '.', ''))
              and a.period_to = (select period_to from latest_agreements)
              and a.amount is not null
        )
        select
            {object_code_sql} as object_code,
            coalesce(nullif(estimate_name, ''), nullif(recipient_name, ''), kcsr_code, {object_code_sql}) as object_name,
            documentclass_id,
            period_to as date_value,
            agreement_number as number_value,
            recipient_name as name_value,
            amount
        from ranked a
        where rn = 1
    """)
    for row in conn.execute(query, {**params, "report_year": REPORT_YEAR}).mappings():
        prefix = "mbt" if row["documentclass_id"] == "273" else "subsidy"
        _merge_detail(target, row["object_code"], prefix, dict(row))


def _fetch_buau(conn, target: dict[str, dict[str, Any]], meta: dict[str, Any], kcsr_code: str | None) -> None:
    mode = meta["mode"]
    params = {"pattern": meta["pattern"], "kcsr_code": kcsr_code}
    object_code_sql = "coalesce(nullif(b.kdr_code, ''), b.kcsr_code)" if mode == "okv" else "replace(b.kcsr_code, '.', '')"
    base_filter = (
        "nullif(b.kdr_code, '') is not null and b.kdr_code not in ('0', '0.0')"
        if mode == "okv"
        else "b.kcsr_code ilike :pattern"
    )
    query = text(f"""
        select
            {object_code_sql} as object_code,
            coalesce(b.organization_name, b.budget_name, b.kcsr_code, {object_code_sql}) as object_name,
            sum(coalesce(b.amount_with_refund, b.amount)) as amount
        from stg.buau_operations b
        where ({base_filter})
          and (:kcsr_code is null or replace(b.kcsr_code, '.', '') = replace(:kcsr_code, '.', ''))
          and extract(year from b.operation_date) = :report_year
          and coalesce(b.amount_with_refund, b.amount) is not null
        group by object_code, object_name
    """)
    for row in conn.execute(query, {**params, "report_year": REPORT_YEAR}).mappings():
        item = target.setdefault(row["object_code"], {
            "object_code": row["object_code"],
            "object_name": row["object_name"] or row["object_code"],
        })
        item["buau_cash_amount"] = item.get("buau_cash_amount", 0.0) + _to_float(row["amount"])


def _load_sheet_data(meta: dict[str, Any], kcsr_code: str | None) -> list[dict[str, Any]]:
    with engine.connect() as conn:
        rows = _fetch_base_rows(conn, meta, kcsr_code)
        _fetch_contracts(conn, rows, meta, kcsr_code)
        _fetch_agreements(conn, rows, meta, kcsr_code)
        _fetch_buau(conn, rows, meta, kcsr_code)
    non_empty_rows = [
        row
        for row in rows.values()
        if any(_to_float(row.get(key)) != 0 for key in AMOUNT_KEYS)
    ]
    return sorted(non_empty_rows, key=lambda row: (row.get("object_name") or "", row.get("object_code") or ""))


def _set_title(ws, title: str, note: str, last_col: int) -> None:
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    ws["A2"] = title
    ws["A3"] = note
    ws["A2"].font = Font(size=13, bold=True)
    ws["A3"].font = Font(size=11, italic=True)


def _merge(ws, cell_range: str, value: str) -> None:
    ws.merge_cells(cell_range)
    ws[cell_range.split(":", 1)[0]].value = value


def _build_headers(ws, meta: dict[str, Any]) -> None:
    last_col = meta["last_col"]
    _set_title(ws, meta["title"], meta["note"], last_col)

    header_ranges = [
        ("A4:A7", meta["name_header"]),
        ("B4:B7", "лимит бюджета субъекта РФ"),
        ("C4:C7", "принятые бюджетные обязательства бюджета субъекта РФ\nвсего"),
        ("D4:H6", "договоры, контракты на закупку товаров, работ, услуг"),
        ("I4:M6", "межбюджетные трансферты местным бюджетам"),
        ("N4:R6", "субсидии бюджетным, автономным учреждениям (ЮЛ, ИП, ФЛ)"),
        ("S4:W6", "договоры, контракты БУ/АУ на закупку товаров, работ, услуг"),
        ("X4:AB6", "кассовые выплаты из бюджета субъекта РФ"),
        ("AC4:AG6", "местные бюджеты: лимиты и обязательства"),
        ("AH4:AL6", "местные бюджеты: контракты и субсидии"),
        ("AM4:AQ6", "кассовые выплаты из местных бюджетов"),
    ]
    if last_col >= 53:
        header_ranges.append(("AR4:BD6", "дополнительные сведения по ОКВ/БУАУ"))

    for cell_range, value in header_ranges:
        _merge(ws, cell_range, value)

    row7 = {
        "D": "итого", "E": "Дата", "F": "Номер", "G": "Контрагент", "H": "Сумма",
        "I": "итого", "J": "Дата", "K": "Номер", "L": "Получатель", "M": "Сумма",
        "N": "итого", "O": "Дата", "P": "Номер", "Q": "Получатель", "R": "Сумма",
        "S": "итого", "T": "Дата", "U": "Номер", "V": "Контрагент", "W": "Сумма",
        "X": "всего", "Y": "контракты", "Z": "субсидии БУ/АУ", "AA": "МБТ", "AB": "касса БУ/АУ",
        "AC": "лимит", "AD": "БО всего", "AE": "контракты", "AF": "субсидии", "AG": "касса БУ/АУ",
        "AH": "Дата", "AI": "Номер", "AJ": "Контрагент", "AK": "Сумма", "AL": "итого БУ/АУ",
        "AM": "всего", "AN": "контракты", "AO": "субсидии", "AP": "касса БУ/АУ", "AQ": "примечание",
        "AR": "факт поставки", "AS": "дата оплаты", "AT": "N документа", "AU": "сумма",
        "AV": "местный лимит", "AW": "местное БО", "AX": "местная касса", "AY": "контракты",
        "AZ": "субсидии", "BA": "касса АУ/БУ", "BB": "код объекта", "BC": "КЦСР", "BD": "ДопКР",
    }
    for column, value in row7.items():
        col_index = column_index_from_string(column)
        if col_index <= last_col:
            ws[column + "7"] = value

    for col in range(1, last_col + 1):
        ws.cell(8, col, col)


def _apply_style(ws, last_col: int) -> None:
    thin = Side(style="thin", color="B8C2D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="DDEBFF")
    subheader_fill = PatternFill("solid", fgColor="EAF2FF")

    widths = {
        "A": 48, "B": 18, "C": 20, "D": 16, "E": 14, "F": 16, "G": 28, "H": 16,
        "I": 16, "J": 14, "K": 16, "L": 28, "M": 16, "N": 16, "O": 14, "P": 16,
        "Q": 28, "R": 16,
    }
    for col in range(1, last_col + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(letter, 15)

    for row in ws.iter_rows(min_row=4, max_row=9, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, size=9)
            cell.fill = header_fill if cell.row in (4, 5, 6) else subheader_fill

    ws.freeze_panes = "A10"
    ws.auto_filter.ref = f"A9:{get_column_letter(last_col)}9"


def _write_value(ws, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row, col, value)
    if isinstance(value, (int, float, Decimal)):
        cell.number_format = NUMBER_FORMAT
    elif hasattr(value, "year"):
        cell.number_format = DATE_FORMAT


def _write_data(ws, rows: list[dict[str, Any]], last_col: int) -> None:
    for offset, item in enumerate(rows):
        row = 10 + offset
        values = {
            1: item.get("object_name"),
            2: item.get("limit_amount"),
            3: item.get("obligation_amount"),
            4: item.get("contract_amount"),
            5: item.get("contract_date"),
            6: item.get("contract_number"),
            7: item.get("contract_name"),
            8: item.get("contract_amount"),
            9: item.get("mbt_amount"),
            10: item.get("mbt_date"),
            11: item.get("mbt_number"),
            12: item.get("mbt_name"),
            13: item.get("mbt_amount"),
            14: item.get("subsidy_amount"),
            15: item.get("subsidy_date"),
            16: item.get("subsidy_number"),
            17: item.get("subsidy_name"),
            18: item.get("subsidy_amount"),
            24: item.get("cash_amount"),
            25: item.get("cash_contract_amount"),
            26: item.get("cash_subsidy_amount"),
            27: item.get("cash_mbt_amount"),
            28: item.get("buau_cash_amount"),
            29: item.get("local_limit_amount"),
            30: item.get("local_obligation_amount"),
            31: item.get("contract_amount"),
            32: item.get("subsidy_amount"),
            33: item.get("buau_cash_amount"),
            34: item.get("contract_date"),
            35: item.get("contract_number"),
            36: item.get("contract_name"),
            37: item.get("contract_amount"),
            38: item.get("subsidy_amount"),
            39: item.get("local_cash_amount"),
            40: item.get("local_cash_contract_amount"),
            41: item.get("local_cash_subsidy_amount"),
            42: item.get("buau_cash_amount"),
            43: item.get("object_code"),
            45: item.get("buau_cash_amount"),
            46: item.get("local_limit_amount"),
            47: item.get("local_obligation_amount"),
            48: item.get("local_cash_amount"),
            49: item.get("local_cash_contract_amount"),
            50: item.get("local_cash_subsidy_amount"),
            51: item.get("buau_cash_amount"),
            52: item.get("object_code"),
            53: item.get("kcsr_code"),
            56: item.get("kdr_code"),
        }
        for col, value in values.items():
            if col <= last_col and value not in (None, ""):
                _write_value(ws, row, col, value)
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.border = Border(
                left=Side(style="thin", color="D6DEE8"),
                right=Side(style="thin", color="D6DEE8"),
                top=Side(style="thin", color="D6DEE8"),
                bottom=Side(style="thin", color="D6DEE8"),
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_analytics_export(section: str | None = None, kcsr_code: str | None = None) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = "Сводная выгрузка nerpochka"
    wb.properties.created = datetime.now()

    for meta in SECTION_SHEETS:
        if section and meta["section"] != section:
            continue
        ws = wb.create_sheet(meta["sheet"])
        _build_headers(ws, meta)
        _apply_style(ws, meta["last_col"])
        _write_data(ws, _load_sheet_data(meta, kcsr_code), meta["last_col"])

    if not wb.worksheets:
        ws = wb.create_sheet("Нет данных")
        ws["A1"] = "Нет данных для выбранных фильтров"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
